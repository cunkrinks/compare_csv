
"""compare_csv.py

Compare two CSV files and export differences using database-style operations.

Usage examples:
  py compare_csv.py a.csv b.csv --key id
  py compare_csv.py a.csv b.csv --key snap,inst --report --excel

Outputs:
  - <prefix>_deleted.csv   (rows only in file A)
  - <prefix>_inserted.csv  (rows only in file B)
  - <prefix>_updated.csv   (rows with same key but different values)

Behavior:
  - Uses larger table as reference automatically
  - Requires both CSVs to have the same set of columns (order may differ)
  - Compares by primary key columns (default: ALL columns)
  - DELETED: rows present in reference but not in compare-to
  - INSERTED: rows present in compare-to but not in reference
  - UPDATED: rows with same key but different non-key values (highlighted in Excel)
"""

import argparse
import os
from typing import List
import pandas as pd
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def read_csv(path: str) -> pd.DataFrame:
    return pd.read_csv(path)


def trim_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Trim whitespace from all string columns."""
    for col in df.columns:
        if df[col].dtype == 'object':
            df[col] = df[col].apply(lambda x: x.strip() if isinstance(x, str) else x)
    return df


def ensure_same_columns(a: pd.DataFrame, b: pd.DataFrame):
    set_a = set(a.columns)
    set_b = set(b.columns)
    if set_a != set_b:
        missing_in_b = sorted(list(set_a - set_b))
        missing_in_a = sorted(list(set_b - set_a))
        msg_parts = []
        if missing_in_b:
            msg_parts.append(f"Columns in A not in B: {missing_in_b}")
        if missing_in_a:
            msg_parts.append(f"Columns in B not in A: {missing_in_a}")
        raise SystemExit("Error: CSV column mismatch. " + "; ".join(msg_parts))


def ensure_outdir(path: str):
    os.makedirs(path, exist_ok=True)


def compare_by_keys(a: pd.DataFrame, b: pd.DataFrame, keys: List[str]):
    """Compare two tables by primary keys (database-style).
    Returns: (deleted, inserted, updated) where updated has suffixes _a/_b for old/new values.
    """
    # DELETED: Rows present in A but keys not in B
    b_keys = b[keys].drop_duplicates()
    a_left = a.merge(b_keys, on=keys, how='left', indicator=True)
    deleted = a_left[a_left['_merge'] == 'left_only'].drop(columns=['_merge'])

    # INSERTED: Rows present in B but keys not in A
    a_keys = a[keys].drop_duplicates()
    b_left = b.merge(a_keys, on=keys, how='left', indicator=True)
    inserted = b_left[b_left['_merge'] == 'left_only'].drop(columns=['_merge'])

    # UPDATED: inner join and compare non-key columns
    inner = a.merge(b, on=keys, how='inner', suffixes=('_a', '_b'))
    non_key_cols = [c for c in a.columns if c not in keys]
    diffs = []
    for col in non_key_cols:
        col_a = f"{col}_a"
        col_b = f"{col}_b"
        if col_a in inner.columns and col_b in inner.columns:
            mask = ~(inner[col_a].fillna('<<NA>>') == inner[col_b].fillna('<<NA>>'))
            if mask.any():
                diffs.append(mask)
    if diffs:
        combined_mask = diffs[0]
        for m in diffs[1:]:
            combined_mask = combined_mask | m
        updated = inner[combined_mask]
    else:
        updated = inner.iloc[0:0]

    return deleted, inserted, updated


def build_combined(a: pd.DataFrame, b: pd.DataFrame, keys: List[str]):
    """Return an outer-joined combined dataframe with status column: SAME/UPDATED/INSERTED/DELETED
    It will include both A and B values with suffixes _a/_b on non-key columns.
    """
    merged = a.merge(b, on=keys, how='outer', indicator=True, suffixes=('_a', '_b'))
    # Determine status
    non_key_cols = [c for c in a.columns if c not in keys]
    def row_status(row):
        if row['_merge'] == 'left_only':
            return 'DELETED'
        if row['_merge'] == 'right_only':
            return 'INSERTED'
        # both -> check diffs
        for col in non_key_cols:
            ca = row.get(col + '_a')
            cb = row.get(col + '_b')
            if (pd.isna(ca) and pd.isna(cb)):
                continue
            if ca != cb:
                return 'UPDATED'
        return 'SAME'
    merged['status'] = merged.apply(row_status, axis=1)
    return merged

def write_df(df: pd.DataFrame, path: str):
    df.to_csv(path, index=False)


def _format_deleted_inserted(df: pd.DataFrame, keys: List[str], source_suffix: str):
    # Return DataFrame with keys + base columns (renaming suffix back to base)
    non_key_cols = [c for c in df.columns if c not in keys and not c.endswith('_a') and not c.endswith('_b') and c not in ['_merge', 'status']]
    # prefer suffix columns
    cols_with_suffix = [k for k in df.columns if k.endswith(source_suffix)]
    base_cols = [c[:-len(source_suffix)] for c in cols_with_suffix]
    out_cols = keys + base_cols
    out_df = pd.DataFrame(columns=out_cols)
    for k in keys:
        out_df[k] = df[k]
    for s_col, bcol in zip(cols_with_suffix, base_cols):
        out_df[bcol] = df[s_col]
    return out_df


def write_excel_sidebyside(a: pd.DataFrame, b: pd.DataFrame, combined: pd.DataFrame, keys: List[str], outpath: str, a_name: str = None, b_name: str = None):
    """Write side-by-side Excel comparison with color-coding.
    - Yellow background for updated rows
    - Empty cells where rows don't exist in the other file
    """
    if not OPENPYXL_AVAILABLE:
        raise SystemExit("Error: --excel flag requires openpyxl. Install with: pip install openpyxl")
    
    wb = openpyxl.Workbook()
    # Create ordered sheets: Summary, Legend, Comparison
    summary_ws = wb.active
    summary_ws.title = "Summary"
    legend_ws = wb.create_sheet(title="Legend")
    ws = wb.create_sheet(title="Comparison")
    
    # Define colors
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    
    # Align columns to the same order (use ordering from a)
    cols_a = list(a.columns)
    cols_b = list(a.columns)  # ensure same order on B side too
    
    # Write headers - use file names if provided
    if a_name:
        a_label = os.path.basename(a_name)
    else:
        a_label = 'Reference (A)'
    if b_name:
        b_label = os.path.basename(b_name)
    else:
        b_label = 'Compare to (B)'

    # Reserve a column between A and B for the Status column
    status_col_index = len(cols_a) + 1
    col_offset_a = 1
    col_offset_status = status_col_index
    col_offset_b = status_col_index + 1

    # Write A header and columns
    ws.cell(row=1, column=col_offset_a, value=f"--- {a_label} ---")
    for i, col in enumerate(cols_a, 1):
        ws.cell(row=2, column=col_offset_a + i - 1, value=col)
    # Write Status header
    ws.cell(row=1, column=col_offset_status, value="--- Status ---")
    ws.cell(row=2, column=col_offset_status, value="Status")
    
    # Write B header and columns
    ws.cell(row=1, column=col_offset_b, value=f"--- {b_label} ---")
    for i, col in enumerate(cols_b, 1):
        ws.cell(row=2, column=col_offset_b + i - 1, value=col)
    
    # Use the provided combined (outer join of A and B with status column)
        # Use the provided combined (outer join of A and B with status column)
    merged = combined
    non_key_cols = [c for c in a.columns if c not in keys]
    def is_row_updated(row):
        if row['_merge'] != 'both':
            return False
        for col in non_key_cols:
            ca = row.get(col + '_a')
            cb = row.get(col + '_b')
            if (pd.isna(ca) and pd.isna(cb)):
                continue
            if ca != cb:
                return True
        return False
    
    # Write rows side-by-side
    row_num = 3
    max_rows = max(len(a), len(b))
    
    # order rows: SAME, UPDATED, DELETED, INSERTED per user request
    same_rows = merged[merged['status'] == 'SAME'] if 'status' in merged.columns else merged.iloc[0:0]
    updated_rows = merged[merged['status'] == 'UPDATED'] if 'status' in merged.columns else merged.iloc[0:0]
    deleted_rows = merged[merged['status'] == 'DELETED'] if 'status' in merged.columns else merged.iloc[0:0]
    inserted_rows = merged[merged['status'] == 'INSERTED'] if 'status' in merged.columns else merged.iloc[0:0]
    # Build combined ordered DataFrame
    ordered = pd.concat([same_rows, updated_rows, deleted_rows, inserted_rows], ignore_index=True)

    for _, row in ordered.iterrows():
        # Build a_row and b_row dicts
        status = row.get('status')
        # For key columns, values are stored without suffix; for non-keys use suffixes
        a_row = {}
        b_row = {}
        for col in cols_a:
            if col in keys:
                a_row[col] = row.get(col)
            else:
                a_row[col] = row.get(col + '_a')
        for col in cols_b:
            if col in keys:
                b_row[col] = row.get(col)
            else:
                b_row[col] = row.get(col + '_b')
        is_updated_row = (status == 'UPDATED')
        
        # Write A columns
        col_num = 1
        # Write A columns - show values only for DELETED, SAME, UPDATED
        if status in ('DELETED', 'SAME', 'UPDATED'):
            for col in cols_a:
                # for SAME and UPDATED use left value; for SAME fallback to right if left missing
                val = a_row.get(col)
                if (pd.isna(val) or val is None) and status == 'SAME':
                    # fallback to right value if left empty for SAME rows
                    val = b_row.get(col)
                if pd.isna(val):
                    val = ''
                cell = ws.cell(row=row_num, column=col_num, value=val)
                # Color left side for DELETED (orange), UPDATED (yellow)
                if status == 'DELETED':
                    cell.fill = orange_fill
                elif is_updated_row:
                    cell.fill = yellow_fill
                col_num += 1
        else:
            # Empty row for A
            for col in cols_a:
                cell = ws.cell(row=row_num, column=col_num, value="")
                col_num += 1
        
        # Write Status column between A and B
        status_col_num = status_col_index
        status_val = status if status is not None else ''
        s_cell = ws.cell(row=row_num, column=status_col_num, value=status_val)
        if status == 'UPDATED':
            s_cell.fill = yellow_fill
        elif status in ('DELETED', 'INSERTED'):
            s_cell.fill = orange_fill

        # Write B columns
        col_num = col_offset_b
        # Write B columns - show values only for INSERTED, SAME, UPDATED
        if status in ('INSERTED', 'SAME', 'UPDATED'):
            for col in cols_b:
                # for SAME and UPDATED use right value; for SAME fallback to left if right missing
                val = b_row.get(col)
                if (pd.isna(val) or val is None) and status == 'SAME':
                    val = a_row.get(col)
                if pd.isna(val):
                    val = ''
                cell = ws.cell(row=row_num, column=col_num, value=val)
                # Color right side for INSERTED (orange), UPDATED (yellow)
                if status == 'INSERTED':
                    cell.fill = orange_fill
                elif is_updated_row:
                    cell.fill = yellow_fill
                col_num += 1
        else:
            # Empty row for B
            for col in cols_b:
                cell = ws.cell(row=row_num, column=col_num, value="")
                col_num += 1
        
        row_num += 1
    
    # Add thin borders for readability across the whole range
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    last_col = col_offset_b + len(cols_b) - 1  # last column index for B side
    # apply border to headers and data
    for r in range(1, row_num):
        for c in range(1, last_col + 1):
            try:
                ws.cell(row=r, column=c).border = border
            except Exception:
                pass

    # Freeze header rows and apply auto-filter for the full table
    ws.freeze_panes = ws['A3']
    last_cell_letter = get_column_letter(last_col)
    # Auto-filter starting from header row 2 to end of data rows
    ws.auto_filter.ref = f"A2:{last_cell_letter}{row_num - 1}"

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
    # Make headers bold and center-align
    header_font = Font(bold=True)
    header_align = Alignment(horizontal='center')
    for c in range(1, last_col + 1):
        try:
            ws.cell(row=1, column=c).font = header_font
            ws.cell(row=1, column=c).alignment = header_align
            ws.cell(row=2, column=c).font = header_font
            ws.cell(row=2, column=c).alignment = header_align
        except Exception:
            pass
    
    # Populate Summary and Legend carefully after table is written
    total_a = len(a)
    total_b = len(b)
    same_count = len(same_rows)
    deleted_count = len(deleted_rows)
    inserted_count = len(inserted_rows)
    updated_count = len(updated_rows)

    # Summary sheet
    summary_ws['A1'] = 'Table Comparison Summary'
    summary_ws['A2'] = 'Reference (A)'
    summary_ws['B2'] = os.path.basename(a_name) if a_name else 'Reference (A)'
    summary_ws['A3'] = 'Compare-to (B)'
    summary_ws['B3'] = os.path.basename(b_name) if b_name else 'Compare-to (B)'
    summary_ws['A5'] = 'Total Rows (A)'
    summary_ws['B5'] = total_a
    summary_ws['A6'] = 'Total Rows (B)'
    summary_ws['B6'] = total_b
    summary_ws['A8'] = 'SAME'
    summary_ws['B8'] = same_count
    summary_ws['A9'] = 'UPDATED'
    summary_ws['B9'] = updated_count
    summary_ws['A10'] = 'DELETED'
    summary_ws['B10'] = deleted_count
    summary_ws['A11'] = 'INSERTED'
    summary_ws['B11'] = inserted_count

    # Legend sheet
    legend_ws['A1'] = 'Legend'
    legend_ws['A3'] = 'UPDATED (Yellow)'
    legend_ws['B3'] = 'Rows with differences in non-key columns between files (A → B)'
    legend_ws['A3'].fill = yellow_fill
    legend_ws['A4'] = 'INSERTED/DELETED (Orange)'
    legend_ws['B4'] = 'Rows present only in one file (left-only or right-only)'
    legend_ws['A4'].fill = orange_fill
    for col in ['A', 'B']:
        legend_ws.column_dimensions[col].width = 40

    wb.save(outpath)


def write_report(deleted: pd.DataFrame, inserted: pd.DataFrame, updated: pd.DataFrame, keys: List[str], outpath: str, total_a: int, total_b: int, same_count: int, a_name: str = None, b_name: str = None):
    """Write a database-style report with DELETED, INSERTED, UPDATED operations."""
    with open(outpath, 'w', encoding='utf-8') as f:
        # summary counts
        f.write(f"\n{'='*80}\n")
        f.write(f"TABLE COMPARISON REPORT (Reference vs Compare-to)\n")
        f.write(f"{'='*80}\n")
        f.write(f"Primary Key Columns: {', '.join(keys) if keys else 'None'}\n")
        f.write(f"Reference (A): {os.path.basename(a_name) if a_name else 'Reference (A)'}\n")
        f.write(f"Compare-to (B): {os.path.basename(b_name) if b_name else 'Compare-to (B)'}\n")
        f.write(f"Total Rows in Reference (A): {total_a}\n")
        f.write(f"Total Rows in Compare-to (B): {total_b}\n")
        f.write(f"Identical Rows: {same_count}\n")
        f.write(f"Deleted Rows: {len(deleted)}\n")
        f.write(f"Inserted Rows: {len(inserted)}\n")
        f.write(f"Updated Rows: {len(updated)}\n")
        f.write(f"{'='*80}\n\n")

        # DELETED rows
        if not deleted.empty:
            f.write(f"\n{'─'*80}\n")
            if a_name:
                f.write(f"ONLY IN {os.path.basename(a_name)}: {len(deleted)} row(s)\n")
            else:
                f.write(f"DELETED: {len(deleted)} row(s) only in Reference (removed from Compare-to)\n")
            f.write(f"{'─'*80}\n")
            for _, row in deleted.iterrows():
                pairs = []
                for col in deleted.columns:
                    val = row[col]
                    if pd.isna(val):
                        sval = 'NULL'
                    else:
                        sval = str(val)
                    pairs.append(f"{col}={sval}")
                line = ' | '.join(pairs)
                f.write(f"{line}\n")

        # INSERTED rows
        if not inserted.empty:
            f.write(f"\n{'─'*80}\n")
            if b_name:
                f.write(f"ONLY IN {os.path.basename(b_name)}: {len(inserted)} row(s)\n")
            else:
                f.write(f"INSERTED: {len(inserted)} row(s) only in Compare-to (new rows)\n")
            f.write(f"{'─'*80}\n")
            for _, row in inserted.iterrows():
                pairs = []
                for col in inserted.columns:
                    val = row[col]
                    if pd.isna(val):
                        sval = 'NULL'
                    else:
                        sval = str(val)
                    pairs.append(f"{col}={sval}")
                line = ' | '.join(pairs)
                f.write(f"{line}\n")

        # UPDATED rows
        if not updated.empty:
            f.write(f"\n{'─'*80}\n")
            f.write(f"DIFFERENT: {len(updated)} row(s) with changed values\n")
            f.write(f"{'─'*80}\n")
            if keys:
                for _, row in updated.iterrows():
                    key_parts = []
                    for k in keys:
                        key_parts.append(f"{k}={row.get(k, 'NULL')}")
                    f.write(f"\nKey: {' | '.join(key_parts)}\n")
                    for col_a in sorted([c for c in updated.columns if c.endswith('_a')]):
                        base = col_a[:-2]
                        col_b = base + '_b'
                        if col_b not in updated.columns:
                            continue
                        val_a = row[col_a]
                        val_b = row[col_b]
                        if pd.isna(val_a) and pd.isna(val_b):
                            continue
                        if (val_a != val_b):
                            a_str = 'NULL' if pd.isna(val_a) else str(val_a)
                            b_str = 'NULL' if pd.isna(val_b) else str(val_b)
                            f.write(f"  {base}: {a_str} → {b_str}\n")


def main():
    p = argparse.ArgumentParser(
        description='Compare two CSV files like database tables (DELETED/INSERTED/UPDATED operations). Larger table is used as reference.',
        epilog='Examples:\n  py compare_csv.py old.csv new.csv --key id\n  py compare_csv.py a.csv b.csv --key snap,inst --report --excel'
    )
    p.add_argument('a', help='First CSV file (will be swapped to be reference if smaller)')
    p.add_argument('b', help='Second CSV file (will be swapped to be reference if smaller)')
    p.add_argument('--key', default=None, help='Comma-separated primary key columns (default: ALL columns)')
    p.add_argument('--outdir', default=None, help='Output directory (default: value of --name)')
    p.add_argument('--name', required=True, help='Name for output files and default outdir')
    p.add_argument('--prefix', default='diff', help='Filename prefix for outputs')
    p.add_argument('--report', action='store_true', help='Write database-style text report (DELETED/INSERTED/UPDATED)')
    p.add_argument('--excel', action='store_true', help='Write side-by-side Excel comparison with yellow highlighting')
    args = p.parse_args()

    a_path = args.a
    b_path = args.b
    outdir = args.outdir or args.name
    prefix = args.prefix

    ensure_outdir(outdir)

    a = read_csv(a_path)
    b = read_csv(b_path)
    
    # Trim whitespace from all string columns
    a = trim_dataframe(a)
    b = trim_dataframe(b)

    # Use larger table as reference (swap if needed)
    if len(b) > len(a):
        print(f"Info: File B is larger ({len(b)} rows) than File A ({len(a)} rows)")
        print(f"Info: Swapping files - B becomes Reference (A)")
        a, b = b, a
        a_path, b_path = b_path, a_path

    # Show short previews of both files before comparing
    print('\n' + '='*60)
    print(f"Reference (A): {a_path}  (rows={a.shape[0]}, cols={a.shape[1]})")
    with pd.option_context('display.max_rows', 10, 'display.max_columns', 20):
        try:
            print(a.head(10).to_string(index=False))
        except Exception:
            print(a.head(10))
    print('\n' + '-'*60)
    print(f"Compare-to (B): {b_path}  (rows={b.shape[0]}, cols={b.shape[1]})")
    with pd.option_context('display.max_rows', 10, 'display.max_columns', 20):
        try:
            print(b.head(10).to_string(index=False))
        except Exception:
            print(b.head(10))
    print('='*60 + '\n')

    # Ensure both files have identical column sets
    ensure_same_columns(a, b)

    # Determine primary key columns
    if args.key:
        keys = [k.strip() for k in args.key.split(',') if k.strip()]
        missing_in_a = [k for k in keys if k not in a.columns]
        missing_in_b = [k for k in keys if k not in b.columns]
        if missing_in_a or missing_in_b:
            raise SystemExit(f"Error: Key columns missing in files: {missing_in_a or missing_in_b}")
    else:
        # Default: use all columns as primary key
        keys = list(a.columns)
        print(f"Info: Using all columns as primary key: {', '.join(keys)}")

    # Sort both frames by keys to make comparisons deterministic
    if keys:
        a = a.sort_values(by=keys, kind='stable').reset_index(drop=True)
        b = b.sort_values(by=keys, kind='stable').reset_index(drop=True)

    # Build merged combined dataframe with status and extract DELETED/INSERTED/UPDATED
    combined = build_combined(a, b, keys)
    deleted = combined[combined['status'] == 'DELETED']
    inserted = combined[combined['status'] == 'INSERTED']
    updated = combined[combined['status'] == 'UPDATED']

    # Write output files
    out_deleted = os.path.join(outdir, f"{prefix}_deleted.csv")
    out_inserted = os.path.join(outdir, f"{prefix}_inserted.csv")
    out_updated = os.path.join(outdir, f"{prefix}_updated.csv")

    # Write clean CSV for DELETED and INSERTED (rename suffixes back)
    deleted_out = _format_deleted_inserted(deleted, keys, '_a') if not deleted.empty else deleted
    inserted_out = _format_deleted_inserted(inserted, keys, '_b') if not inserted.empty else inserted
    write_df(deleted_out, out_deleted)
    write_df(inserted_out, out_inserted)
    write_df(updated, out_updated)
    # Write the combined merged CSV with status
    out_combined = os.path.join(outdir, f"{prefix}_combined.csv")
    write_df(combined, out_combined)

    # Optional database-style text report
    if getattr(args, 'report', False):
        out_report = os.path.join(outdir, f"{prefix}_report.txt")
        total_a = len(a)
        total_b = len(b)
        # Compute identical rows count: total_a minus deleted minus updated
        same_count = max(0, total_a - len(deleted) - len(updated))
        write_report(deleted, inserted, updated, keys, out_report, total_a, total_b, same_count, a_path, b_path)
        print(f"Wrote report: {out_report}")

    # Optional Excel export
    if getattr(args, 'excel', False):
        out_excel = os.path.join(outdir, f"{prefix}_comparison.xlsx")
        write_excel_sidebyside(a, b, combined, keys, out_excel, a_path, b_path)
        print(f"Wrote Excel: {out_excel}")

    print(f"Wrote: {out_deleted} ({len(deleted)} DELETED rows)")
    print(f"Wrote: {out_inserted} ({len(inserted)} INSERTED rows)")
    print(f"Wrote: {out_updated} ({len(updated)} UPDATED rows)")


if __name__ == '__main__':
    main()
