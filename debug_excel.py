import openpyxl
wb=openpyxl.load_workbook('compare_object/diff_comparison.xlsx')
print(wb.sheetnames)
ws=wb['Comparison']
for r in range(1,6):
    row=[ws.cell(row=r,column=c).value for c in range(1,20)]
    print(row)
print('\nSummary:')
summary=wb['Summary']
for r in range(1,15):
    print(r, [summary.cell(row=r,column=c).value for c in range(1,3)])
print('\nLegend entries:')
legend=wb['Legend']
for r in range(1,6):
    print(r, [legend.cell(row=r,column=c).value for c in range(1,3)])
print('\nFreeze panes:', ws.freeze_panes)
print('Auto-filter ref:', ws.auto_filter.ref)
